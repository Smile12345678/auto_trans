import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelWriter():

	def __init__(self):
		pass

	def setLangData(self, langData):
		self.m_langData = langData

	def writeLangTable(self, index):
		i18n_path = ""
		lang_name = ""
		if index == 1:
			i18n_path = "lang_zh.xls"
			lang_name = "zh"
		elif index == 2:
			i18n_path = "lang_en.xls"
			lang_name = "en"
		elif index == 3:
			i18n_path = "lang_jp.xls"
			lang_name = "jp"

		wb = Workbook()
		sheet = wb.create_sheet("lang",0)
		sheet.cell(row = 1,column= 1).value="id"
		sheet.cell(row = 1,column= 2).value="raw"
		sheet.cell(row = 1,column= 3).value=lang_name
		writeindex = 2

		for item in self.m_langData:
			a = item
			id = item[0]
			en = item[1]
			langData = item[2]
			sheet.cell(row = writeindex,column= 1).value=id
			sheet.cell(row = writeindex,column= 2).value=en
			sheet.cell(row = writeindex,column= 3).value=langData
			writeindex = writeindex + 1

		wb.save(i18n_path)



